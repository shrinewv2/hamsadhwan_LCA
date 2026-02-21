"""Excel Agent — parse Excel/CSV files via E2B sandbox execution."""
import io
import json
from typing import Any, Dict, List

from backend.agents.base_agent import BaseAgent
from backend.models.schemas import FileMetadata, ParsedOutput
from backend.processing.bedrock_client import invoke_claude_haiku, parse_json_response
from backend.processing.e2b_sandbox import execute_excel_analysis
from backend.utils.logger import append_job_log, get_logger

logger = get_logger("excel_agent")

LCA_KEYWORDS = [
    "Impact Category", "CO2", "GWP", "Functional Unit", "Process",
    "Ecoinvent", "CO2 eq", "kg CO2", "MJ", "LCA", "emission",
    "inventory", "impact", "characterisation", "normalisation",
    "global warming", "acidification", "eutrophication", "ozone",
    "ecotoxicity", "human toxicity", "land use", "water",
]


class ExcelAgent(BaseAgent):
    agent_name = "excel_agent"

    def process(self, file_meta: FileMetadata, file_bytes: bytes) -> ParsedOutput:
        # Step 1 — Structure Inspection (no sandbox)
        sheet_names, row_counts, sample_headers = self._inspect_structure(file_bytes, file_meta)

        append_job_log(
            file_meta.job_id, "INFO", self.agent_name, file_meta.file_id,
            f"Structure: {len(sheet_names)} sheets, headers sampled"
        )

        # Step 2 — Code generation via Claude Haiku
        code = self._generate_analysis_code(sheet_names, row_counts, sample_headers)

        # Step 3 — Sandbox execution
        exit_code, stdout, stderr = execute_excel_analysis(
            file_bytes, code, filename="input_file"
        )

        if exit_code == 0 and stdout:
            try:
                parsed = json.loads(stdout)
                markdown = self._build_markdown(parsed)
                return ParsedOutput(
                    file_id=file_meta.file_id,
                    job_id=file_meta.job_id,
                    agent=self.agent_name,
                    markdown=markdown,
                    structured_json=parsed,
                    lca_relevant=parsed.get("lca_data_found", False),
                    confidence=0.95,
                    word_count=len(markdown.split()),
                )
            except json.JSONDecodeError:
                append_job_log(
                    file_meta.job_id, "WARN", self.agent_name, file_meta.file_id,
                    "Sandbox output not valid JSON, trying simpler code"
                )

        # Retry Attempt 2 — Simpler code
        simple_code = self._generate_simple_code()
        exit_code, stdout, stderr = execute_excel_analysis(
            file_bytes, simple_code, filename="input_file"
        )

        if exit_code == 0 and stdout:
            try:
                parsed = json.loads(stdout)
                markdown = self._build_markdown(parsed)
                return ParsedOutput(
                    file_id=file_meta.file_id,
                    job_id=file_meta.job_id,
                    agent=self.agent_name,
                    markdown=markdown,
                    structured_json=parsed,
                    lca_relevant=parsed.get("lca_data_found", False),
                    confidence=0.80,
                    word_count=len(markdown.split()),
                )
            except json.JSONDecodeError:
                pass

        # Retry Attempt 3 — Local openpyxl fallback
        append_job_log(
            file_meta.job_id, "WARN", self.agent_name, file_meta.file_id,
            "Falling back to local openpyxl parsing"
        )
        return self._openpyxl_fallback(file_meta, file_bytes)

    def _inspect_structure(self, file_bytes: bytes, file_meta: FileMetadata):
        """Step 1: Inspect Excel structure locally with openpyxl."""
        sheet_names = []
        row_counts = {}
        sample_headers = {}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            sheet_names = wb.sheetnames
            for name in sheet_names:
                ws = wb[name]
                row_counts[name] = ws.max_row or 0
                # Sample first row as headers
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(c) if c is not None else "" for c in row]
                sample_headers[name] = headers[:20]  # Cap at 20 columns
            wb.close()
        except Exception as e:
            logger.warning("structure_inspection_failed", error=str(e))
            # Try CSV fallback
            if file_meta.file_type.value == "CSV":
                try:
                    import csv
                    text = file_bytes.decode("utf-8", errors="ignore")
                    reader = csv.reader(io.StringIO(text))
                    first_row = next(reader, [])
                    sheet_names = ["Sheet1"]
                    row_counts = {"Sheet1": sum(1 for _ in reader) + 1}
                    sample_headers = {"Sheet1": first_row[:20]}
                except Exception:
                    pass

        return sheet_names, row_counts, sample_headers

    def _generate_analysis_code(self, sheet_names: list, row_counts: dict, sample_headers: dict) -> str:
        """Step 2: Use Claude Haiku to generate pandas analysis code."""
        prompt = f"""Generate a complete Python script that:
1. Reads the Excel file at '/home/user/input_file' using pandas
2. The file has these sheets: {sheet_names}
3. Row counts per sheet: {row_counts}
4. Sample headers per sheet: {json.dumps(sample_headers)}
5. For each non-empty sheet:
   - Convert to a Markdown table using df.to_markdown()
   - Check columns for LCA keywords: {LCA_KEYWORDS}
   - If LCA columns found, flag lca_relevant=True and extract summary stats
6. Output a single JSON object to stdout with keys:
   - sheets: list of {{name, markdown, lca_relevant, columns}}
   - lca_data_found: bool
   - errors: list of strings
7. Contains NO network calls, NO shell commands, NO file system access beyond reading /home/user/input_file
8. Import pandas, json, sys at the top

Return ONLY the Python code block. No explanation, no preamble."""

        system = "You are a Python code generator. Return ONLY executable Python code, no markdown fences."
        response = invoke_claude_haiku(prompt, system)
        # Strip any markdown fences
        code = response.strip()
        if code.startswith("```python"):
            code = code[9:]
        elif code.startswith("```"):
            code = code[3:]
        if code.endswith("```"):
            code = code[:-3]
        return code.strip()

    def _generate_simple_code(self) -> str:
        """Attempt 2: Generate simpler fallback code."""
        return '''
import pandas as pd
import json
import sys

try:
    sheets_data = pd.read_excel("/home/user/input_file", sheet_name=None)
    result = {"sheets": [], "lca_data_found": False, "errors": []}
    
    lca_keywords = ["impact", "co2", "gwp", "emission", "lca", "energy", "kg", "mj"]
    
    for name, df in sheets_data.items():
        if df.empty:
            continue
        try:
            md = df.to_markdown(index=False)
        except Exception:
            md = df.to_string()
        
        cols = [str(c).lower() for c in df.columns]
        lca_rel = any(kw in " ".join(cols) for kw in lca_keywords)
        if lca_rel:
            result["lca_data_found"] = True
        
        result["sheets"].append({
            "name": str(name),
            "markdown": md,
            "lca_relevant": lca_rel,
            "columns": list(df.columns.astype(str))
        })
    
    print(json.dumps(result))
except Exception as e:
    print(json.dumps({"sheets": [], "lca_data_found": False, "errors": [str(e)]}))
'''

    def _openpyxl_fallback(self, file_meta: FileMetadata, file_bytes: bytes) -> ParsedOutput:
        """Attempt 3: Local openpyxl-based parsing."""
        markdown_parts = []
        structured = {"sheets": [], "lca_data_found": False, "errors": []}
        lca_found = False

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

            for name in wb.sheetnames:
                ws = wb[name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])

                if not rows or all(all(c == "" for c in row) for row in rows):
                    continue

                # Build Markdown table
                md_lines = [f"## Sheet: {name}\n"]
                if rows:
                    header = rows[0]
                    md_lines.append("| " + " | ".join(header) + " |")
                    md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
                    for row in rows[1:]:
                        # Pad or trim row to match header length
                        padded = row + [""] * (len(header) - len(row))
                        md_lines.append("| " + " | ".join(padded[:len(header)]) + " |")

                md = "\n".join(md_lines)
                markdown_parts.append(md)

                # Check for LCA relevance
                all_text = " ".join(str(c).lower() for row in rows for c in row)
                lca_rel = any(kw.lower() in all_text for kw in LCA_KEYWORDS)
                if lca_rel:
                    lca_found = True

                structured["sheets"].append({
                    "name": name,
                    "markdown": md,
                    "lca_relevant": lca_rel,
                    "columns": rows[0] if rows else [],
                })

            wb.close()
        except Exception as e:
            structured["errors"].append(str(e))

        structured["lca_data_found"] = lca_found
        full_md = "\n\n".join(markdown_parts) if markdown_parts else "# No data extracted"

        return ParsedOutput(
            file_id=file_meta.file_id,
            job_id=file_meta.job_id,
            agent=self.agent_name,
            markdown=full_md,
            structured_json=structured,
            lca_relevant=lca_found,
            confidence=0.70,
            word_count=len(full_md.split()),
        )

    def _build_markdown(self, parsed: dict) -> str:
        """Convert parsed JSON into a Markdown document."""
        parts = []
        for sheet in parsed.get("sheets", []):
            name = sheet.get("name", "Unknown")
            md = sheet.get("markdown", "")
            lca = " *(LCA relevant)*" if sheet.get("lca_relevant") else ""
            parts.append(f"## Sheet: {name}{lca}\n\n{md}")
        return "\n\n".join(parts) if parts else "# No data extracted"
