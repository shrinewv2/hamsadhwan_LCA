"""FastAPI router for /upload endpoints — Ingestion Service."""
import uuid
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, UploadFile

from backend.config import settings
from backend.ingestion.complexity_scorer import compute_complexity_score, estimate_processing_seconds
from backend.ingestion.file_detector import detect_file_type, probe_pdf_structure
from backend.ingestion.virus_scanner import scan_file
from backend.models.enums import FileStatus, FileType
from backend.models.schemas import (
    AnalysisRecord,
    ErrorResponse,
    FileMetadata,
    JobCreateResponse,
)
from backend.storage import dynamo_client, s3_client
from backend.utils.logger import append_job_log, get_logger, init_job_log_buffer

logger = get_logger("ingestion")

router = APIRouter(prefix="/api/v1", tags=["ingestion"])


async def _process_job(job_id: str, file_metas: List[FileMetadata]):
    """Background task: trigger the orchestrator to process the job."""
    try:
        from backend.orchestrator.graph import run_pipeline
        await run_pipeline(job_id, file_metas)
    except Exception as e:
        logger.error("pipeline_failed", job_id=job_id, error=str(e))
        append_job_log(job_id, "ERROR", "orchestrator", None, f"Pipeline failed: {str(e)}")
        try:
            dynamo_client.update_analysis_status(job_id, "FAILED")
        except Exception:
            pass


@router.post("/jobs", response_model=JobCreateResponse, responses={400: {"model": ErrorResponse}})
async def create_job(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    user_context: Optional[str] = Form(None),
):
    """
    Upload one or more LCA documents and start analysis.
    Accepts multipart/form-data with 1–20 files.
    """
    # Validate file count
    max_files = settings.MAX_FILES_PER_JOB if settings else 20
    if len(files) < 1:
        raise HTTPException(status_code=400, detail="At least 1 file is required")
    if len(files) > max_files:
        raise HTTPException(status_code=400, detail=f"Maximum {max_files} files per job")

    job_id = str(uuid.uuid4())
    init_job_log_buffer(job_id)
    append_job_log(job_id, "INFO", "ingestion", None, f"Job {job_id} created with {len(files)} files")

    max_size = settings.max_file_size_bytes if settings else 100 * 1024 * 1024
    file_metas: List[FileMetadata] = []
    total_estimated_seconds = 0

    # Parse user_context JSON if provided
    import json
    user_ctx = {}
    if user_context:
        try:
            user_ctx = json.loads(user_context)
        except json.JSONDecodeError:
            user_ctx = {"raw_context": user_context}

    for upload_file in files:
        file_bytes = await upload_file.read()

        # Validate file size
        if len(file_bytes) > max_size:
            raise HTTPException(
                status_code=400,
                detail=f"File '{upload_file.filename}' exceeds maximum size of {settings.MAX_FILE_SIZE_MB if settings else 100} MB",
            )

        # Virus scan
        scan_result = scan_file(file_bytes, upload_file.filename or "")
        if not scan_result["clean"]:
            raise HTTPException(
                status_code=400,
                detail=f"File '{upload_file.filename}' failed virus scan: {scan_result['details']}",
            )

        # Detect file type via magic bytes
        file_type, detected_mime = detect_file_type(file_bytes, upload_file.filename or "")

        # PDF structure probing
        pdf_info = {}
        if file_type == FileType.PDF:
            pdf_info = probe_pdf_structure(file_bytes)

        # Excel structure probing for complexity
        sheet_count = 0
        estimated_row_count = 0
        if file_type in (FileType.EXCEL,):
            try:
                import openpyxl
                import io
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
                sheet_count = len(wb.sheetnames)
                for ws in wb:
                    estimated_row_count += ws.max_row or 0
                wb.close()
            except Exception:
                sheet_count = 1
                estimated_row_count = 100

        # Compute complexity
        complexity = compute_complexity_score(
            file_type=file_type,
            page_count=pdf_info.get("page_count", 0),
            has_embedded_images=pdf_info.get("has_embedded_images", False),
            is_scanned=pdf_info.get("is_scanned", False),
            sheet_count=sheet_count,
            estimated_row_count=estimated_row_count,
            size_bytes=len(file_bytes),
        )

        file_id = str(uuid.uuid4())
        s3_key = f"uploads/{job_id}/{file_id}/{upload_file.filename}"

        # Upload to S3
        try:
            s3_client.upload_file_bytes(
                bucket=settings.S3_BUCKET_UPLOADS if settings else "lca-uploads",
                key=s3_key,
                data=file_bytes,
                content_type=detected_mime,
            )
        except Exception as e:
            logger.warning("s3_upload_skipped", error=str(e))
            # Continue without S3 for local dev

        # Build metadata record
        meta = FileMetadata(
            file_id=file_id,
            job_id=job_id,
            original_name=upload_file.filename or "unknown",
            s3_key=s3_key,
            actual_mime=detected_mime,
            file_type=file_type,
            size_bytes=len(file_bytes),
            is_scanned=pdf_info.get("is_scanned", False),
            has_text_layer=pdf_info.get("has_text_layer", False),
            has_embedded_images=pdf_info.get("has_embedded_images", False),
            page_count=pdf_info.get("page_count", None),
            sheet_count=sheet_count if sheet_count > 0 else None,
            complexity_score=complexity,
            status=FileStatus.PENDING,
        )

        # Write to DynamoDB
        try:
            dynamo_client.put_file_record(meta.model_dump())
        except Exception as e:
            logger.warning("dynamo_write_skipped", error=str(e))

        file_metas.append(meta)
        total_estimated_seconds += estimate_processing_seconds(complexity)

        append_job_log(
            job_id, "INFO", "ingestion", file_id,
            f"Registered file: {upload_file.filename} (type={file_type.value}, complexity={complexity})"
        )

    # Create analysis record
    analysis = AnalysisRecord(
        job_id=job_id,
        status="PENDING",
        file_ids=[m.file_id for m in file_metas],
        user_context=user_ctx,
    )
    try:
        dynamo_client.put_analysis_record(analysis.model_dump())
    except Exception as e:
        logger.warning("analysis_record_write_skipped", error=str(e))

    # Schedule background processing
    background_tasks.add_task(_process_job, job_id, file_metas)

    return JobCreateResponse(
        job_id=job_id,
        file_count=len(file_metas),
        estimated_seconds=total_estimated_seconds,
        status="PENDING",
    )
